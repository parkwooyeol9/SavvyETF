"""Export ADR analysis to Excel with aligned data and embedded charts."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from adr_data_loader import EVENT_BUFFER_DAYS

PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_DIR / "data" / "adr_output"
KST = ZoneInfo("Asia/Seoul")

CHART_IMAGE_WIDTH = 920
CHART_IMAGE_HEIGHT = 520
CHART_ROW_SPAN = 28


def _format_event_sheet(event: pd.DataFrame) -> pd.DataFrame:
    out = event.reset_index().rename(columns={"index": "date", "Date": "date"})
    cols = [
        c
        for c in [
            "date",
            "close",
            "volume",
            "daily_return",
            "days_from_listing",
            "trading_day_offset",
            "phase",
            "price_index",
            "rebased_return_pct",
        ]
        if c in out.columns
    ]
    return out[cols]


def build_listing_dates_dataframe(analysis: dict) -> pd.DataFrame:
    rows: list[dict] = []
    for result in analysis.get("results", []):
        m = result["metrics"]
        rows.append(
            {
                "adr_symbol": m["adr_symbol"],
                "underlying_symbol": m["underlying_symbol"],
                "company_name": m["company_name"],
                "us_adr_listing_date": m["us_adr_listing_date"],
                "analysis_event_date": m["analysis_event_date"],
                "analysis_event_source": m["analysis_event_source"],
                "data_source": m.get("data_source", ""),
            }
        )
    return pd.DataFrame(rows)


def build_aligned_compare_dataframe(analysis: dict) -> pd.DataFrame:
    """
    Wide table for t=0 aligned comparison.

    Columns: trading_day_offset, {ADR}_return_pct, ...
    trading_day_offset = 0 is the first trading day on/after ADR listing.
    """
    merged: pd.DataFrame | None = None
    for result in analysis.get("results", []):
        m = result["metrics"]
        sym = m["adr_symbol"]
        event = result["event"]
        if "trading_day_offset" not in event.columns or "rebased_return_pct" not in event.columns:
            continue

        part = (
            event[["trading_day_offset", "rebased_return_pct"]]
            .drop_duplicates(subset=["trading_day_offset"])
            .rename(columns={"rebased_return_pct": f"{sym}_return_pct"})
            .set_index("trading_day_offset")
            .sort_index()
        )
        merged = part if merged is None else merged.join(part, how="outer")

    if merged is None:
        return pd.DataFrame(columns=["trading_day_offset"])
    return merged.reset_index().sort_values("trading_day_offset")


def _add_aligned_line_chart(workbook, sheet_name: str = "Aligned_Compare") -> None:
    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 2:
        return

    chart = LineChart()
    chart.title = "Aligned cumulative return (rebased at t=0)"
    chart.y_axis.title = "Return (%)"
    chart.x_axis.title = "Trading-day offset (t=0 = ADR listing)"
    chart.width = 24
    chart.height = 12
    chart.legend.position = "b"

    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    anchor_col = get_column_letter(max_col + 2)
    ws.add_chart(chart, f"{anchor_col}2")


def _add_image_block(ws, row: int, title: str, image_path: Path) -> int:
    ws.cell(row=row, column=1, value=title)
    img = XLImage(str(image_path))
    img.width = CHART_IMAGE_WIDTH
    img.height = CHART_IMAGE_HEIGHT
    ws.add_image(img, f"A{row + 1}")
    return row + CHART_ROW_SPAN


def _add_charts_sheet(workbook, chart_paths: dict[str, Path]) -> None:
    if not chart_paths:
        return

    if "Charts" in workbook.sheetnames:
        del workbook["Charts"]
    ws = workbook.create_sheet("Charts")

    row = 1
    ordered: list[tuple[str, str]] = [
        ("Summary panel (pre vs post)", "summary_panel"),
        ("Aligned overlay — t=0 rebased returns", "aligned_overlay"),
    ]
    used: set[str] = set()
    for title, key in ordered:
        path = chart_paths.get(key)
        if path and path.exists():
            row = _add_image_block(ws, row, title, path)
            used.add(key)

    for key, path in sorted(chart_paths.items()):
        if key in used or not path.exists():
            continue
        row = _add_image_block(ws, row, f"{key} — event chart (listing date marked)", path)


def export_analysis_excel(
    analysis: dict,
    run_id: str | None = None,
    chart_paths: dict[str, Path] | None = None,
) -> Path:
    run_id = run_id or datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    out_dir = OUTPUT_DIR / run_id
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"adr_impact_{run_id}.xlsx"

    summary = analysis["summary"].copy()
    listing_dates = build_listing_dates_dataframe(analysis)
    aligned_compare = build_aligned_compare_dataframe(analysis)

    notes = pd.DataFrame(
        {
            "field": [
                "generated_at",
                "window",
                "pre_post_definition",
                "listing_date_source",
                "significance_test",
                "aligned_compare_sheet",
                "charts_sheet",
            ],
            "value": [
                datetime.now(KST).strftime("%Y-%m-%d %H:%M KST"),
                "±2 calendar years around analysis event date",
                f"Pre/post exclude ±{EVENT_BUFFER_DAYS} days around event (buffer)",
                "US ADR listing date from registry/yfinance; underlying prices from Finnhub/FinMind/EODHD/Yahoo",
                "Welch t-test: post vs pre daily returns (skipped if limited pre data)",
                "Aligned_Compare: trading_day_offset with rebased_return_pct per ADR (0% at t=0)",
                "Charts: embedded PNGs + native line chart on Aligned_Compare",
            ],
        }
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        notes.to_excel(writer, sheet_name="README", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        listing_dates.to_excel(writer, sheet_name="Listing_Dates", index=False)
        aligned_compare.to_excel(writer, sheet_name="Aligned_Compare", index=False)

        for result in analysis["results"]:
            sym = result["metrics"]["adr_symbol"]
            sheet = sym[:31]
            _format_event_sheet(result["event"]).to_excel(writer, sheet_name=sheet, index=False)

        if analysis.get("errors"):
            pd.DataFrame({"errors": analysis["errors"]}).to_excel(
                writer, sheet_name="Errors", index=False
            )

        workbook = writer.book
        _add_aligned_line_chart(workbook)
        if chart_paths:
            _add_charts_sheet(workbook, chart_paths)

    return path
